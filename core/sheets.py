"""数据表格分析 —— CSV / xlsx 解析 + 统计 + 预览。

设计取舍：
- CSV 用标准库 csv 解析，零依赖；xlsx 用 openpyxl（可选，缺则给提示）。
- 解析后在内存保留结构化数据，前端拿 preview + stats 渲染，
  喂给 AI 时只注入「列统计 + 前若干行」文本上下文（控制 token 量）。
- 不做重的聚合计算，复杂统计交给 AI 在上下文里判断或给处理建议。
"""

from __future__ import annotations

import csv
import os
import statistics
from typing import Any

PREVIEW_ROWS = 30
MAX_CELLS_CONTEXT = 4000  # 注入 AI 上下文的单元格上限


def _csv_cell(v: Any) -> str:
    s = "" if v is None else str(v)
    if any(ch in s for ch in (",", '"', "\n")):
        s = '"' + s.replace('"', '""') + '"'
    return s


class Table:
    def __init__(self, name: str, headers: list[str], rows: list[list[Any]]):
        self.name = name
        self.headers = headers
        self.rows = rows

    @property
    def n_rows(self) -> int:
        return len(self.rows)

    @property
    def n_cols(self) -> int:
        return len(self.headers)

    def column_stats(self) -> list[dict[str, Any]]:
        stats: list[dict[str, Any]] = []
        for ci, h in enumerate(self.headers):
            col_vals = [r[ci] if ci < len(r) else None for r in self.rows]
            non_null = [v for v in col_vals if v not in (None, "")]
            nums: list[float] = []
            is_numeric = True
            for v in non_null:
                try:
                    nums.append(float(v))
                except (TypeError, ValueError):
                    is_numeric = False
                    break
            nulls = sum(1 for v in col_vals if v in (None, ""))
            entry: dict[str, Any] = {
                "name": h,
                "non_null": len(non_null),
                "nulls": nulls,
                "type": "numeric" if (is_numeric and nums) else "text",
            }
            if is_numeric and nums:
                entry["min"] = min(nums)
                entry["max"] = max(nums)
                entry["mean"] = round(statistics.fmean(nums), 4)
            stats.append(entry)
        return stats

    def preview(self, n: int = PREVIEW_ROWS) -> list[dict[str, Any]]:
        out: list[dict[str, Any]] = []
        for r in self.rows[:n]:
            out.append(
                {self.headers[i]: (r[i] if i < len(r) else "") for i in range(len(self.headers))}
            )
        return out

    def to_context(self, max_cells: int = MAX_CELLS_CONTEXT) -> str:
        """生成给 AI 的文本上下文：列信息 + 统计 + 前若干行预览。"""
        lines = [f"表格：{self.name}", f"行数：{self.n_rows}，列数：{self.n_cols}"]
        stats = self.column_stats()
        lines.append("列信息：")
        for s in stats:
            if s["type"] == "numeric":
                lines.append(
                    f"  - {s['name']}（数值型，非空 {s['non_null']}，空 {s['nulls']}，"
                    f"最小 {s['min']}，最大 {s['max']}，均值 {s['mean']}）"
                )
            else:
                lines.append(f"  - {s['name']}（文本型，非空 {s['non_null']}，空 {s['nulls']}）")
        lines.append("")
        lines.append("数据预览（前若干行，CSV 格式）：")
        lines.append(",".join(str(h) for h in self.headers))
        cells = 0
        for r in self.rows:
            line = ",".join(_csv_cell(r[i] if i < len(r) else "") for i in range(len(self.headers)))
            lines.append(line)
            cells += len(self.headers)
            if cells >= max_cells:
                lines.append("...（预览截断）")
                break
        return "\n".join(lines)


def load_csv(path: str) -> Table:
    with open(path, "r", encoding="utf-8-sig", errors="ignore", newline="") as f:
        rows_raw = list(csv.reader(f))
    if not rows_raw:
        return Table(os.path.basename(path), [], [])
    headers = rows_raw[0]
    data = rows_raw[1:]
    return Table(os.path.basename(path), headers, data)


def load_xlsx(path: str) -> Table:
    from openpyxl import load_workbook  # type: ignore

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_raw = list(ws.iter_rows(values_only=True))
    if not rows_raw:
        return Table(os.path.basename(path), [], [])
    headers = [str(c) if c is not None else "" for c in rows_raw[0]]
    data = [[c if c is not None else "" for c in row] for row in rows_raw[1:]]
    return Table(os.path.basename(path), headers, data)


def load_table(path: str) -> Table:
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xls"):
        return load_xlsx(path)
    return load_csv(path)
